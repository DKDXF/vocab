"""
单词/词书相关 API
"""
import csv
import io
import uuid
from fastapi import APIRouter, HTTPException, UploadFile, File
from database import get_db, close_db
from models import BookOut, WordOut, WordWithRootOut, BookProgressOut, ImportResult

router = APIRouter(prefix="/api/books", tags=["单词/词书"])


@router.get("", response_model=list[BookOut], summary="获取词书列表及进度")
def get_books():
    """获取所有词书，附带学习进度信息"""
    conn = get_db()
    try:
        current_row = conn.execute(
            "SELECT value FROM settings WHERE key = 'current_book_id'"
        ).fetchone()
        current_book_id = current_row["value"] if current_row else "cet4"

        books = conn.execute("SELECT * FROM word_books ORDER BY id").fetchall()
        result = []
        for b in books:
            total = b["word_count"]
            progress_rows = conn.execute(
                """SELECT status, COUNT(*) as cnt FROM word_progress wp
                   JOIN words w ON wp.word_id = w.id
                   WHERE w.book_id = ?
                   GROUP BY status""",
                (b["id"],),
            ).fetchall()
            status_map = {r["status"]: r["cnt"] for r in progress_rows}
            learned = sum(v for k, v in status_map.items() if k != "new")
            mastered = status_map.get("mastered", 0)
            pct = round((learned / total * 100), 1) if total > 0 else 0.0

            # 双记忆率
            memory_rows = conn.execute(
                """SELECT wp.history
                   FROM word_progress wp
                   JOIN words w ON wp.word_id = w.id
                   WHERE w.book_id = ? AND wp.history != ''""",
                (b["id"],),
            ).fetchall()
            total_h_len = 0; total_h_zero = 0; total_r_len = 0; total_r_zero = 0
            for r in memory_rows:
                h = r["history"]
                total_h_len += len(h); total_h_zero += h.count('0')
                recent = h[-2:] if len(h) >= 2 else h
                total_r_len += len(recent); total_r_zero += recent.count('0')
            history_rate = round((1 - total_h_zero / total_h_len) * 100, 1) if total_h_len > 0 else 0.0
            recent_rate = round((1 - total_r_zero / total_r_len) * 100, 1) if total_r_len > 0 else 0.0

            result.append(BookOut(
                id=b["id"],
                name=b["name"],
                description=b["description"],
                icon=b["icon"],
                word_count=total,
                learned=learned,
                mastered=mastered,
                progress_pct=pct,
                is_current=(b["id"] == current_book_id),
                history_rate=history_rate,
                recent_rate=recent_rate,
            ))
        return result
    finally:
        close_db(conn)


@router.get("/{book_id}/words", response_model=list[WordWithRootOut], summary="获取词书的单词列表")
def get_book_words(book_id: str):
    """获取指定词书的所有单词（含词根信息）"""
    conn = get_db()
    try:
        book = conn.execute("SELECT id FROM word_books WHERE id = ?", (book_id,)).fetchone()
        if not book:
            raise HTTPException(status_code=404, detail="词书不存在")

        rows = conn.execute(
            """SELECT w.*, wr.root_text, wr.meaning as root_meaning
               FROM words w
               LEFT JOIN word_roots wr ON w.root_id = wr.id
               WHERE w.book_id = ? ORDER BY w.sort_order""",
            (book_id,),
        ).fetchall()
        return [WordWithRootOut(
            id=r["id"], book_id=r["book_id"], word=r["word"],
            phonetic=r["phonetic"], part_of_speech=r["part_of_speech"],
            definition_cn=r["definition_cn"], example_sentence=r["example_sentence"],
            example_translation=r["example_translation"],
            root_id=r["root_id"], high_freq_defs=r["high_freq_defs"],
            confusion_group=r["confusion_group"],
            root_text=r["root_text"] or "", root_meaning=r["root_meaning"] or "",
            mnemonic=r["mnemonic"] if "mnemonic" in r.keys() else "",
            synonym=r["synonym"] if "synonym" in r.keys() else "",
            antonym=r["antonym"] if "antonym" in r.keys() else "",
            derivative=r["derivative"] if "derivative" in r.keys() else "",
            note=r["note"] if "note" in r.keys() else "",
        ) for r in rows]
    finally:
        close_db(conn)


@router.post("/{book_id}/select", summary="选择当前词书")
def select_book(book_id: str):
    """设置当前使用的词书"""
    conn = get_db()
    try:
        book = conn.execute("SELECT id FROM word_books WHERE id = ?", (book_id,)).fetchone()
        if not book:
            raise HTTPException(status_code=404, detail="词书不存在")

        conn.execute(
            "INSERT INTO settings (key, value) VALUES ('current_book_id', ?) "
            "ON CONFLICT(key) DO UPDATE SET value = ?",
            (book_id, book_id),
        )
        conn.commit()
        return {"message": f"已切换到 {book_id}", "current_book_id": book_id}
    finally:
        close_db(conn)


@router.delete("/{book_id}", summary="删除词书")
def delete_book(book_id: str):
    """删除指定词书及其所有单词、学习进度"""
    conn = get_db()
    try:
        # 检查词书是否存在
        book = conn.execute("SELECT id, name FROM word_books WHERE id = ?", (book_id,)).fetchone()
        if not book:
            raise HTTPException(status_code=404, detail="词书不存在")
        
        # 不允许删除系统默认词书（可选）
        if book_id in ["cet4", "cet6", "kaoyan", "ielts", "toefl"]:
            raise HTTPException(status_code=403, detail="系统默认词书不可删除")
        
        # 如果删除的是当前词书，需要重置为默认词书
        current_row = conn.execute(
            "SELECT value FROM settings WHERE key = 'current_book_id'"
        ).fetchone()
        if current_row and current_row["value"] == book_id:
            conn.execute(
                "INSERT INTO settings (key, value) VALUES ('current_book_id', ?) "
                "ON CONFLICT(key) DO UPDATE SET value = ?",
                ("cet4", "cet4"),
            )
        
        # 删除词书（级联删除相关数据）
        # 1. 删除单词进度
        conn.execute(
            """DELETE FROM word_progress 
               WHERE word_id IN (SELECT id FROM words WHERE book_id = ?)""",
            (book_id,),
        )
        # 2. 删除学习记录
        conn.execute(
            """DELETE FROM study_records 
               WHERE word_id IN (SELECT id FROM words WHERE book_id = ?)""",
            (book_id,),
        )
        # 3. 删除单词
        conn.execute("DELETE FROM words WHERE book_id = ?", (book_id,))
        # 4. 删除词书
        conn.execute("DELETE FROM word_books WHERE id = ?", (book_id,))
        
        conn.commit()
        return {"message": f"已删除词书: {book['name']}", "deleted_book_id": book_id}
    finally:
        close_db(conn)


@router.post("/import", summary="导入词书（支持 Excel/CSV）")
async def import_book(file: UploadFile = File(...), book_name: str = None):
    """
    上传 .xlsx/.xls/.csv 文件导入新词书
    表头必须包含 word 列，可选列：phonetic, part_of_speech, definition_cn,
    example_sentence, example_translation, root, high_freq_def
    """
    filename = file.filename or ""
    ext = filename.rsplit(".", 1)[-1].lower() if "." in filename else ""

    if ext not in ("xlsx", "xls", "csv"):
        raise HTTPException(status_code=400, detail="仅支持 .xlsx/.xls/.csv 文件")

    content = await file.read()

    try:
        if ext == "csv":
            rows_data = _parse_csv(content)
        else:
            rows_data = _parse_excel(content)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"文件解析失败: {str(e)}")

    if not rows_data:
        raise HTTPException(status_code=400, detail="文件为空或格式不正确")

    # 校验表头
    headers = list(rows_data[0].keys())
    if "word" not in headers:
        raise HTTPException(status_code=400, detail="表头必须包含 'word' 列")

    # 生成词书ID和名称
    if not book_name or not book_name.strip():
        book_name = filename.rsplit(".", 1)[0]
    book_id = f"import_{uuid.uuid4().hex[:8]}"

    conn = get_db()
    try:
        # 创建词书
        conn.execute(
            "INSERT INTO word_books (id, name, description, icon, word_count) VALUES (?, ?, ?, ?, ?)",
            (book_id, book_name, f"导入的词书: {filename}", "📥", len(rows_data)),
        )

        success_count = 0
        fail_count = 0
        errors = []

        # 处理词根映射
        root_cache = {}

        for idx, row in enumerate(rows_data):
            word = (row.get("word") or "").strip()
            if not word:
                fail_count += 1
                errors.append(f"第{idx+2}行: word 为空")
                continue

            phonetic = (row.get("phonetic") or "").strip()
            part_of_speech = (row.get("part_of_speech") or "").strip()
            definition_cn = (row.get("definition_cn") or "").strip()
            example_sentence = (row.get("example_sentence") or "").strip()
            example_translation = (row.get("example_translation") or "").strip()
            root_text = (row.get("root") or "").strip()
            high_freq_def = (row.get("high_freq_def") or "").strip()

            # 处理词根
            root_id = None
            if root_text:
                if root_text not in root_cache:
                    existing = conn.execute(
                        "SELECT id FROM word_roots WHERE root_text = ?", (root_text,)
                    ).fetchone()
                    if existing:
                        root_cache[root_text] = existing["id"]
                    else:
                        cursor = conn.execute(
                            "INSERT INTO word_roots (root_text, meaning, description) VALUES (?, ?, ?)",
                            (root_text, "", f"从导入文件自动创建"),
                        )
                        root_cache[root_text] = cursor.lastrowid
                root_id = root_cache[root_text]

            try:
                conn.execute(
                    """INSERT INTO words
                       (book_id, word, phonetic, part_of_speech, definition_cn,
                        example_sentence, example_translation, sort_order, root_id, high_freq_defs)
                       VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
                    (book_id, word, phonetic, part_of_speech, definition_cn,
                     example_sentence, example_translation, idx, root_id, high_freq_def),
                )
                success_count += 1
            except Exception as e:
                fail_count += 1
                errors.append(f"第{idx+2}行 '{word}': {str(e)}")

        conn.commit()

        return ImportResult(
            success_count=success_count,
            fail_count=fail_count,
            errors=errors[:20],  # 最多返回20条错误
            book_id=book_id,
            book_name=book_name,
        )
    finally:
        close_db(conn)


def _parse_csv(content: bytes) -> list:
    """解析 CSV 文件"""
    text = content.decode("utf-8-sig")
    reader = csv.DictReader(io.StringIO(text))
    return [row for row in reader]


def _parse_excel(content: bytes) -> list:
    """解析 Excel 文件"""
    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(content), read_only=True)
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    headers = [str(h or "").strip().lower() for h in next(rows_iter)]
    result = []
    for row in rows_iter:
        d = {}
        for i, h in enumerate(headers):
            if h and i < len(row):
                d[h] = str(row[i] or "")
        result.append(d)
    wb.close()
    return result
